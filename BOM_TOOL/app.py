# 必要なライブラリをインポート
from flask import Flask, request, jsonify, render_template
import openpyxl
import re
import io
import csv
import pdfplumber
import json
import traceback
# zipfile と xml は openpyxl(rich_text=True) を使うため不要になりました

# Flaskアプリケーションを作成
app = Flask(__name__)

# --- 正規表現とキーワード ---
ref_pattern = re.compile(r'[A-Z]+[0-9]+')
# Q5-8 と Q5-Q8 の両方を正しく処理できるように修正
ref_range_pattern = re.compile(r'^([A-Z]+)(\d+)\s*[-~～]\s*([A-Z]*)(\d+)$', re.IGNORECASE)

HEADER_KEYWORDS = {
    'ref': ['部品番号', 'ref des', 'ロケーション番号', 'ref', '記号', 'designator', 'symbol', 'リファレンス', '回路記号', '位置番号', '部品記号'],
    'part': ['part number', 'メーカー品番', '型番', '型式', '形式', '型格', '定格', 'part', 'value', '品名', 'description', '図番', '名称', 'パート名'],
    'mfg': ['メーカー', 'mfg', 'maker', 'manufacturer', '製造元']
}

# --- 型番からメーカーを推測する関数 ---
def detect_manufacturer(part_number_string):
    pn_upper = part_number_string.upper()
    if pn_upper.startswith(('GRM', 'GCM', 'BLM')): return 'Murata'
    if pn_upper.startswith('CGA'): return 'TDK'
    if pn_upper.startswith('MCR'): return 'Rohm'
    if pn_upper.startswith('CC'): return 'Yageo'
    pn_lower = part_number_string.lower()
    if 'murata' in pn_lower: return 'Murata'
    if 'tdk' in pn_lower: return 'TDK'
    if 'rohm' in pn_lower: return 'Rohm'
    if 'yageo' in pn_lower: return 'Yageo'
    if 'kyocera' in pn_lower: return 'Kyocera'
    return ""

# --- 1. Webページの表示 ---
@app.route('/')
def index():
    return render_template('index.html')

# --- 2. ファイル処理のエンドポイント ---
@app.route('/process', methods=['POST'])
def process_file_endpoint():
    if 'file' not in request.files: return jsonify({"error": "ファイルがありません"}), 400
    file = request.files['file']
    if file.filename == '': return jsonify({"error": "ファイルが選択されていません"}), 400
    
    filename = file.filename.lower()
    in_memory_file = io.BytesIO(file.read())
    
    all_flat_data = []
    individual_results = {}
    
    try:
        if filename.endswith(('.xlsx', '.xls')):
            selected_sheets_json = request.form.get('sheets', '[]')
            selected_sheets = json.loads(selected_sheets_json)
            
            if not selected_sheets:
                return jsonify({"error": "処理するシートが選択されていません。"}), 400

            # ### ▼▼▼ 修正箇所 ▼▼▼ ###
            # zipfile方式をやめ、rich_text=True を使う方式に戻します
            try:
                workbook = openpyxl.load_workbook(in_memory_file, rich_text=True)
            except Exception as e:
                print(traceback.format_exc())
                return jsonify({"error": f"Excelファイルの読み込みに失敗しました。サポートされている .xlsx 形式か確認してください。 (エラー: {e})"}), 500
            # ### ▲▲▲ 修正完了 ▲▲▲ ###
            
            for sheet_name in selected_sheets:
                if sheet_name not in workbook.sheetnames:
                    individual_results[sheet_name] = {"error": "指定されたシートが見つかりません。"}
                    continue
                
                sheet = workbook[sheet_name]
                
                # ### ▼▼▼ 修正箇所 ▼▼▼ ###
                # rich_text 対応のパーサーを呼び出します
                data_2d, cancellation_refs = parse_single_excel_sheet_rich_text(sheet)
                # ### ▲▲▲ 修正完了 ▲▲▲ ###
                
                # コアロジックを呼び出し（フラットリストを取得）
                flat_list, error = extract_flat_list_from_rows(data_2d, cancellation_refs)
                
                if error:
                    individual_results[sheet_name] = {"error": error}
                else:
                    individual_results[sheet_name] = group_and_finalize_bom(flat_list)
                    all_flat_data.extend(flat_list)

        else:
            # Excel以外のファイル（PDF, CSV, TXT）の処理
            data_2d, cancellation_refs = [], set()
            if filename.endswith('.csv'):
                data_2d = parse_csv_or_txt(in_memory_file, delimiters=[','])
            elif filename.endswith('.txt'):
                data_2d = parse_csv_or_txt(in_memory_file, delimiters=['\t', r'\s{2,}'])
            elif filename.endswith('.pdf'):
                data_2d = parse_pdf(in_memory_file)
            else:
                return jsonify({"error": "対応していないファイル形式です。"}), 400

            if not data_2d: return jsonify({"error": "ファイルからデータを抽出できませんでした。"}), 500
            
            flat_list, error = extract_flat_list_from_rows(data_2d, cancellation_refs)
            if error: return jsonify({"error": error}), 500
            
            all_flat_data.extend(flat_list)
            individual_results = {}

        combined_results = group_and_finalize_bom(all_flat_data)
        
        if not combined_results:
             return jsonify({"error": "有効なデータが見つかりませんでした。"}), 500

        return jsonify({
            "combined": combined_results,
            "individual": individual_results
        })
        
    except Exception as e:
        print(traceback.format_exc())
        return jsonify({"error": f"処理中に予期せぬエラーが発生しました: {e}"}), 500


# --- 3. 各ファイル形式のパーサー ---

# ### ▼▼▼ 修正箇所 (rich_text対応のExcelパーサー) ▼▼▼ ###
# rich_text=True モードで読み込んだセルを処理する
def parse_single_excel_sheet_rich_text(sheet):
    data = []
    cancellation_refs = set()
    
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            cell_full_text = ""
            
            if cell.value is None:
                row_data.append("")
                continue

            if isinstance(cell.value, list):
                # ■ It's Rich Text (a list of TextRun or str objects)
                
                text_to_cancel = ""
                
                for run in cell.value:
                    # --- ここからが修正点 ---
                    if isinstance(run, str):
                        # run が 'str' オブジェクトの場合
                        cell_full_text += run
                    elif hasattr(run, 'text'):
                        # run が TextRun オブジェクトの場合
                        if run.text:
                            cell_full_text += run.text
                            # Check for partial strikethrough
                            if run.font and run.font.strike:
                                # 取り消し線のテキストをバッファに追加
                                text_to_cancel += " " + run.text
                    # --- ここまでが修正点 ---
                
                row_data.append(cell_full_text)
                
                # セル内の全ての run を確認した後、
                # 取り消し線テキスト全体から部品番号を抽出
                if text_to_cancel:
                    found_refs = ref_pattern.findall(text_to_cancel)
                    for ref in found_refs:
                        cancellation_refs.add(ref)
            
            else:
                # ■ It's a simple value (string, number, etc.)
                cell_full_text = str(cell.value)
                row_data.append(cell_full_text)
                
                # Check for cell-level strikethrough
                if cell.font and cell.font.strike:
                    found_refs = ref_pattern.findall(cell_full_text)
                    for ref in found_refs:
                        cancellation_refs.add(ref)
        
        data.append(row_data)
        
    return data, cancellation_refs
# ### ▲▲▲ 修正完了 ▲▲▲ ###


def parse_csv_or_txt(file_stream, delimiters):
    file_stream.seek(0)
    try: text_data = file_stream.read().decode('utf-8')
    except UnicodeDecodeError:
        file_stream.seek(0)
        text_data = file_stream.read().decode('shift_jis', errors='replace')
    lines = text_data.splitlines()
    data_2d = []
    if len(delimiters) == 1: # CSV
        reader = csv.reader(lines)
        for row in reader:
            cleaned_row = [cell.strip().strip('"').strip(',').strip() for cell in row]
            data_2d.append(cleaned_row)
    else: # TXT
        delimiter_regex = '|'.join(delimiters)
        for line in lines:
            split_row = re.split(delimiter_regex, line)
            cleaned_row = []
            for cell in split_row:
                cleaned_cell = cell.strip().strip('"').strip(',').strip()
                cleaned_row.append(cleaned_cell)
            data_2d.append(cleaned_row)
    return data_2d

def parse_pdf(file_stream):
    data_2d = []
    with pdfplumber.open(file_stream) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table: data_2d.extend(table)
            else:
                text = page.extract_text()
                if text:
                    for line in text.split('\n'): data_2d.append(re.split(r'\s{2,}', line))
    cleaned_data_2d = []
    for row in data_2d:
        if isinstance(row, list):
             cleaned_row = [str(cell).strip().strip('"').strip(',').strip() if cell is not None else "" for cell in row]
             cleaned_data_2d.append(cleaned_row)
    return cleaned_data_2d

# --- 4. コアロジック ---

def extract_flat_list_from_rows(data_2d, cancellation_refs=set()):
    header_map, header_row_index, best_score = {}, -1, 0
    for i, row in enumerate(data_2d[:20]):
        if not isinstance(row, list): continue
        temp_map, used_cols = {}, set()
        for key in ['ref', 'part', 'mfg']:
            for keyword in HEADER_KEYWORDS[key]:
                found = False # ### 「↑」バグ修正のため、ここのロジックを元に戻します
                for j, cell in enumerate(row):
                    if j in used_cols: continue
                    if keyword.replace(" ", "") in str(cell).lower().strip().replace(" ", ""):
                        temp_map[key] = j; used_cols.add(j); found = True; break
                if found: break
        score = len(temp_map)
        if score > best_score:
            best_score, header_map, header_row_index = score, temp_map, i
            if best_score == 3: break
    if best_score < 2: return None, "ヘッダー行（「部品番号」と「型番」など）の特定に失敗しました。"

    flat_list, last_valid = [], {}
    start_index = header_row_index + 1 if header_row_index != -1 else 0
    current_refs_from_last_row = []

    for row in data_2d[start_index:]:
        if not isinstance(row, list) or all(c is None or str(c).strip() == "" for c in row): continue
        def get_cell_value(key):
            idx = header_map.get(key)
            return str(row[idx]).strip() if idx is not None and len(row) > idx and row[idx] is not None else ""

        ref_val_raw = get_cell_value('ref')
        part_val_raw = get_cell_value('part')
        mfg_val_raw = get_cell_value('mfg')
        
        # ### ▼▼▼ 修正箇所 (「↑」機能のバグ修正) ▼▼▼ ###
        is_part_continuation = part_val_raw in ['上↑', '↑', '"']
        is_mfg_continuation = mfg_val_raw in ['上↑', '↑', '"']
        
        if is_part_continuation: part_val_raw = last_valid.get('part', '')
        elif part_val_raw: last_valid['part'] = part_val_raw
        if is_mfg_continuation: mfg_val_raw = last_valid.get('mfg', '')
        elif mfg_val_raw: last_valid['mfg'] = mfg_val_raw

        ref_val = ref_val_raw.replace('(', '').replace(')', '').replace('（', '').replace('）', '')

        if ref_val:
            # この行に新しい Ref がある場合
            potential_parts = [r for r in re.split(r'[,、\s]+', ref_val) if r]
            expanded_refs = []
            for part in potential_parts:
                range_match = ref_range_pattern.match(part)
                if range_match:
                    prefix, start, opt_prefix, end = range_match.groups()
                    if start and end:
                        try:
                            if not opt_prefix or prefix.upper() == opt_prefix.upper():
                                for i in range(int(start), int(end) + 1): 
                                    expanded_refs.append(f"{prefix}{i}")
                            else:
                                expanded_refs.append(part)
                        except ValueError:
                            expanded_refs.append(part)
                    else:
                         expanded_refs.append(part)
                else:
                    expanded_refs.append(part)
            
            # 除外リストを適用
            current_refs_from_last_row = [r for r in expanded_refs if r and r not in cancellation_refs]
        
        elif not is_part_continuation and not is_mfg_continuation:
            # この行に Ref がなく、かつ「上↑」でもない場合 (つまり新しい部品)
            # 誤って前の行の Ref を引き継がないようにリセットする
            current_refs_from_last_row = []
        # ### ▲▲▲ 修正完了 ▲▲▲ ###


        part_val_list = [p.strip() for p in part_val_raw.split('\n') if p.strip()]

        if not any(part_val_list) or not current_refs_from_last_row:
            continue

        for part_line in part_val_list:
            part_val = part_line.split()[0] if part_line else ""
            mfg_val = mfg_val_raw if mfg_val_raw else detect_manufacturer(part_val)
            if part_val:
                for r in current_refs_from_last_row:
                    flat_list.append({"ref": r, "part": part_val, "mfg": mfg_val})
    
    return flat_list, None


def group_and_finalize_bom(flat_list):
    grouped_map = {}
    for item in flat_list:
        key = f"{item['part']}||{item['mfg']}"
        if key not in grouped_map:
            grouped_map[key] = {'refs': set(), 'part': item['part'], 'mfg': item['mfg']}
        grouped_map[key]['refs'].add(item['ref'])

    final_results = []
    for group in grouped_map.values():
        sorted_refs = sorted(list(group['refs']), key=lambda x: [int(t) if t.isdigit() else t.lower() for t in re.split('([0-9]+)', x)])
        final_results.append({'ref': ', '.join(sorted_refs), 'part': group['part'], 'mfg': group['mfg']})

    return final_results


# --- 5. サーバーの起動 ---
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)